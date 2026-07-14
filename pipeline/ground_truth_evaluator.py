#!/usr/bin/env python3
"""Evaluate extracted outputs against chemistry ground truth."""

from __future__ import annotations

import argparse
import csv
import json
import re
from dataclasses import dataclass
from pathlib import Path
from difflib import SequenceMatcher
from typing import Dict, List, Tuple, Any

import openpyxl


DEFAULT_PAPER_MAP = {
    "polymerPaper1": "223",
    "polymerPaper2": "251",
    "polymerPaper3": "124",
}


@dataclass
class FieldEval:
    status: str
    coverage: float
    matched: int
    expected: int
    ground_truth: str
    extracted: str


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = text.replace("°", " ")
    text = text.replace("μ", "u")
    text = re.sub(r"\s+", " ", text)
    text = re.sub(r"[^a-z0-9./+\-\s]", "", text)
    return text.strip()


def split_multi(value: Any) -> List[str]:
    text = str(value or "").strip()
    if not text:
        return []
    parts = re.split(r",|;|/|\band\b", text, flags=re.IGNORECASE)
    cleaned = [normalize_text(p) for p in parts if normalize_text(p)]
    return cleaned


def text_match(expected: str, candidate: str) -> bool:
    if not expected or not candidate:
        return False
    if expected in candidate or candidate in expected:
        return True
    ratio = SequenceMatcher(a=expected, b=candidate).ratio()
    return ratio >= 0.78


def load_ground_truth(xlsx_path: Path) -> Dict[str, Dict[str, Any]]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    idx = {str(h).strip(): i + 1 for i, h in enumerate(headers) if h}

    rows: Dict[str, Dict[str, Any]] = {}
    for r in range(2, ws.max_row + 1):
        ref = ws.cell(r, idx["Reference"]).value
        if ref is None:
            continue
        ref_key = str(int(ref)) if isinstance(ref, (int, float)) else str(ref).strip()

        polymer = ws.cell(r, idx["Polymer"]).value
        alt_polymer = ws.cell(r, idx["Alternate Polymer Names"]).value
        solvents = ws.cell(r, idx["Solvents"]).value
        stationary_phase = ws.cell(r, idx["Stationary Phase"]).value
        detector = ws.cell(r, idx["Detector"]).value
        temperature = ws.cell(r, idx["Temperature (Celsius)"]).value
        flow_rate = ws.cell(r, idx["Flow Rate (mL/min)"]).value
        critical_low = ws.cell(r, idx["Critical Range Low (kD)"]).value
        critical_high = ws.cell(r, idx["Critical Range High (kD)"]).value

        polymers = [normalize_text(polymer), normalize_text(alt_polymer)]
        polymers = [p for p in polymers if p]

        rows[ref_key] = {
            "reference": ref_key,
            "polymer_names": polymers,
            "solvents": split_multi(solvents),
            "stationary_phase": normalize_text(stationary_phase),
            "detectors": split_multi(detector),
            "temperature": normalize_text(temperature),
            "flow_rate": normalize_text(flow_rate),
            "critical_range": normalize_text(
                f"{critical_low}-{critical_high}" if critical_low is not None and critical_high is not None else ""
            ),
        }

    return rows


def load_extracted_json(results_dir: Path, paper_stem: str) -> Dict[str, Any]:
    path = results_dir / f"{paper_stem}_latest.json"
    if not path.exists():
        raise FileNotFoundError(f"Missing extracted file: {path}")
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def evaluate_set(
    expected_values: List[str],
    candidate_values: List[str],
    ground_truth_label: str,
    extracted_label: str,
) -> FieldEval:
    expected_values = [normalize_text(v) for v in expected_values if normalize_text(v)]
    candidate_values = [normalize_text(v) for v in candidate_values if normalize_text(v)]

    if not expected_values:
        return FieldEval("out_of_scope", 1.0, 0, 0, ground_truth_label, extracted_label)

    matched = 0
    for exp in expected_values:
        if any(text_match(exp, cand) for cand in candidate_values):
            matched += 1

    coverage = matched / len(expected_values)
    if coverage == 1:
        status = "exact"
    elif coverage > 0:
        status = "partial"
    else:
        status = "missing"

    return FieldEval(status, coverage, matched, len(expected_values), ground_truth_label, extracted_label)


def evaluate_scalar(
    expected_value: str,
    candidate_values: List[str],
    ground_truth_label: str,
    extracted_label: str,
) -> FieldEval:
    expected_value = normalize_text(expected_value)
    candidate_values = [normalize_text(v) for v in candidate_values if normalize_text(v)]
    if not expected_value:
        return FieldEval("out_of_scope", 1.0, 0, 0, ground_truth_label, extracted_label)

    matched = 1 if any(text_match(expected_value, c) for c in candidate_values) else 0
    coverage = float(matched)
    status = "exact" if matched else "missing"
    return FieldEval(status, coverage, matched, 1, ground_truth_label, extracted_label)


def evaluate_one_paper(
    paper_stem: str,
    ref_key: str,
    gt: Dict[str, Any],
    extracted_json: Dict[str, Any],
) -> Dict[str, Any]:
    ex = extracted_json.get("extracted_data", {})
    master_rows = ex.get("master_table", [])
    metadata_rows = ex.get("column_system_metadata", [])

    polymers = []
    for row in master_rows:
        polymers.extend([row.get("polymer_system", ""), row.get("target_at_cc", "")])

    detections = []  # current three-table schema does not include explicit detector rows

    extracted_stationary = [r.get("stationary_phase", "") for r in master_rows] + [
        r.get("stationary_phase_chemistry", "") for r in metadata_rows
    ]
    extracted_mobile = [r.get("mobile_phase", "") for r in master_rows]
    extracted_temp = [r.get("temp_c", "") for r in master_rows]
    extracted_flow = []  # out of scope in current table schema
    extracted_mw_range = []  # out of scope in current table schema

    field_results = {
        "polymer_names": evaluate_set(
            gt["polymer_names"],
            polymers,
            ", ".join(gt["polymer_names"]),
            ", ".join([p for p in polymers if p]),
        ),
        "detectors": evaluate_set(
            gt["detectors"],
            detections,
            ", ".join(gt["detectors"]),
            ", ".join([d for d in detections if d]),
        ),
        "stationary_phase": evaluate_scalar(
            gt["stationary_phase"],
            extracted_stationary,
            gt["stationary_phase"],
            ", ".join([s for s in extracted_stationary if s]),
        ),
        "solvents_mobile_phase": evaluate_set(
            gt["solvents"],
            extracted_mobile,
            ", ".join(gt["solvents"]),
            ", ".join([m for m in extracted_mobile if m]),
        ),
        "temperature": evaluate_scalar(
            gt["temperature"],
            extracted_temp,
            gt["temperature"],
            ", ".join([t for t in extracted_temp if t]),
        ),
        "flow_rate": evaluate_scalar(
            gt["flow_rate"],
            extracted_flow,
            gt["flow_rate"],
            ", ".join([f for f in extracted_flow if f]),
        ),
        "critical_range": evaluate_scalar(
            gt["critical_range"],
            extracted_mw_range,
            gt["critical_range"],
            ", ".join([r for r in extracted_mw_range if r]),
        ),
    }

    score_points = 0.0
    score_max = 0.0
    for result in field_results.values():
        if result.status == "out_of_scope":
            continue
        score_max += 1
        if result.status == "exact":
            score_points += 1
        elif result.status == "partial":
            score_points += 0.5

    paper_score = (score_points / score_max) if score_max else 0.0

    return {
        "paper": paper_stem,
        "reference": ref_key,
        "score": round(paper_score, 4),
        "field_results": {
            name: {
                "status": r.status,
                "coverage": round(r.coverage, 4),
                "matched": r.matched,
                "expected": r.expected,
                "ground_truth": r.ground_truth,
                "extracted": r.extracted,
            }
            for name, r in field_results.items()
        },
    }


def write_csv_report(results: List[Dict[str, Any]], output_csv: Path) -> None:
    output_csv.parent.mkdir(parents=True, exist_ok=True)
    with open(output_csv, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "paper",
                "reference",
                "field",
                "status",
                "coverage",
                "matched",
                "expected",
                "ground_truth",
                "extracted",
            ],
        )
        writer.writeheader()
        for paper_result in results:
            for field, field_result in paper_result["field_results"].items():
                writer.writerow(
                    {
                        "paper": paper_result["paper"],
                        "reference": paper_result["reference"],
                        "field": field,
                        "status": field_result["status"],
                        "coverage": field_result["coverage"],
                        "matched": field_result["matched"],
                        "expected": field_result["expected"],
                        "ground_truth": field_result["ground_truth"],
                        "extracted": field_result["extracted"],
                    }
                )


def summarize(results: List[Dict[str, Any]]) -> Dict[str, Any]:
    counts = {"exact": 0, "partial": 0, "missing": 0, "out_of_scope": 0}
    for paper_result in results:
        for field_result in paper_result["field_results"].values():
            status = field_result["status"]
            counts[status] = counts.get(status, 0) + 1

    avg_score = 0.0
    if results:
        avg_score = sum(r["score"] for r in results) / len(results)

    return {
        "papers_evaluated": len(results),
        "average_score": round(avg_score, 4),
        "field_status_counts": counts,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Evaluate extracted outputs against chemistry ground truth.")
    parser.add_argument(
        "--ground-truth",
        default="Docs/Papers-Example-PolyCrit.xlsx",
        help="Path to ground-truth XLSX file",
    )
    parser.add_argument(
        "--results-dir",
        default="results",
        help="Directory containing <paper>_latest.json files",
    )
    parser.add_argument(
        "--output-dir",
        default="evaluation",
        help="Directory for evaluation outputs",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    ground_truth_path = Path(args.ground_truth)
    results_dir = Path(args.results_dir)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    gt_rows = load_ground_truth(ground_truth_path)

    paper_results = []
    for paper_stem, ref_key in DEFAULT_PAPER_MAP.items():
        if ref_key not in gt_rows:
            continue
        extracted = load_extracted_json(results_dir, paper_stem)
        paper_results.append(
            evaluate_one_paper(
                paper_stem=paper_stem,
                ref_key=ref_key,
                gt=gt_rows[ref_key],
                extracted_json=extracted,
            )
        )

    summary = summarize(paper_results)
    report = {
        "summary": summary,
        "paper_results": paper_results,
        "mapping": DEFAULT_PAPER_MAP,
        "notes": [
            "This is v0 schema-aligned evaluation on overlapping fields only.",
            "Statuses: exact=1.0, partial=0.5, missing=0.0.",
            "Out-of-scope fields are excluded from paper score.",
        ],
    }

    json_path = output_dir / "ground_truth_evaluation.json"
    csv_path = output_dir / "ground_truth_field_scores.csv"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(report, f, indent=2, ensure_ascii=False)
    write_csv_report(paper_results, csv_path)

    print(f"Saved JSON report: {json_path}")
    print(f"Saved CSV report: {csv_path}")
    print(f"Average score: {summary['average_score']}")
    print(f"Field counts: {summary['field_status_counts']}")


if __name__ == "__main__":
    main()
